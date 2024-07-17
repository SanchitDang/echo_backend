import json
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib.auth import login
from django.contrib import messages
from django.http import HttpResponseRedirect
from django.http import JsonResponse
from apis.models import Users, Bids, Unit
from .models import PanelUser,Assessments,Banner
from .forms import LoginForm,UserForm,AssessmentForm,ItemsSubCategoriesForm,ItemsCategoryForm,MyprofleForm,DomainsForm,UserTypeForm,ProductForm,BannerForm,UnitForm,PanelUserForm
from apis.models import  ItemsSubCategories,ItemsCategory,Scraps,Services,Products,Domains,UserType,Refers
import os
import firebase_admin
from firebase_admin import credentials
from firebase_admin import firestore
import logging
import openpyxl
from django.http import HttpResponse


parent_directory = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
firebase_json_path = os.path.join(parent_directory, 'firebase.json')
cred = credentials.Certificate(firebase_json_path)
firebase_admin.initialize_app(cred)

def update_current_bid(group_id, new_bid):
    # Initialize Firestore client
    db = firestore.Client()

    try:
        logging.info(f"Updating current bid for group ID: {group_id}, New bid: {new_bid}")

        # Query the Firestore collection to find the first document where 'bidId' matches 'group_id'
        query = db.collection('groups').where('bidId', '==', int(group_id)).limit(1)
        docs = query.stream()

        # Update only the first document found (if any)
        for doc in docs:
            logging.info("Document found. Updating currentBid field.")
            # Get the reference to the document
            group_ref = doc.reference

            # Update the currentBid field of the document
            group_ref.update({'currentBid': int(new_bid)})
            
            return JsonResponse({'success': True, 'message': f'Current bid updated to {new_bid} for group {group_id}'})
        
        # If no document is found with matching bidId, return error response
        logging.warning(f"No document found with bidId {group_id}")
        return JsonResponse({'success': False, 'message': f'No document found with bidId {group_id}'})
    
    except Exception as e:
        logging.error(f"Error occurred: {e}")
        return JsonResponse({'success': False, 'message': str(e)})


def panel_login(request):
    
    form = LoginForm()

    if request.method == "POST":
        form = LoginForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']
            user_type = form.cleaned_data['user_type']
            user = PanelUser.objects.filter(email=email, password=password).first()
            if user is not None:
                login(request, user,backend='home.models.PanelUserAuthBackend')

                user_type = request.user.user_type
                if user_type == 'Admin':
                    return redirect('dashboard')
                elif user_type == 'Execution':
                    return redirect('execution_dashboard')
                elif user_type == 'Service_support':
                    return redirect('service_support_dashboard')
                elif user_type == 'Freelancers':
                    return redirect('freelancers_dashboard')
                elif user_type == 'referral':
                    return redirect('referral_dashboard')
                
                
            else:
                return render(request, 'login.html', {'form': form, 'error': 'Invalid credentials'})
        else:
            return render(request, 'login.html', {'form': form, 'error': 'Invalid form data'})
        
    return render(request, 'login.html', {'form': form})



def panel_logout(request):

    return redirect('panel_login')

@login_required(login_url='panel_login')
def dashboard(request):
    # Get the count of all users
    user_count = Users.objects.count()

    # Get the count of users based on user_type
    supplier_count = Users.objects.filter(user_type='supplier').count()
    manufacturer_count = Users.objects.filter(user_type='manufacturer').count()
    referral_count = Users.objects.filter(user_type='referral').count()
    service_provider_count = Users.objects.filter(user_type='service provider').count()

    bid_count = Bids.objects.count()
    on_going_bid_count= Bids.objects.filter(bid_status='on_going').count()
    bid_finished_count = Bids.objects.filter(bid_status='finished').count()
    bid_cancelled_count = Bids.objects.filter(bid_status='cancelled').count()


    context = {
        "manufacturer": "manufacturer",
        "supplier": "supplier",
        'user_count': user_count,
        'bid_count': bid_count,
        'supplier_count': supplier_count,
        'service_provider_count':service_provider_count,
        'manufacturer_count': manufacturer_count,
        'referral_count': referral_count,
        "bid_cancelled_count": bid_cancelled_count,
        "on_going_bid_count": on_going_bid_count,
        "bid_finished_count": bid_finished_count
    }

   

    return render(request, 'dashboard.html', context)


@login_required(login_url='panel_login')
def execution_dashboard(request):
    # Get the count of all users
    user_count = Users.objects.count()

    # Get the count of users based on user_type
    supplier_count = Users.objects.filter(user_type='supplier').count()
    manufacturer_count = Users.objects.filter(user_type='manufacturer').count()
    referral_count = Users.objects.filter(user_type='referral').count()
    service_provider_count = Users.objects.filter(user_type='service provider').count()

    bid_count = Bids.objects.count()
    on_going_bid_count= Bids.objects.filter(bid_status='on_going').count()
    bid_finished_count = Bids.objects.filter(bid_status='finished').count()
    bid_cancelled_count = Bids.objects.filter(bid_status='cancelled').count()


    context = {
        "manufacturer": "manufacturer",
        "supplier": "supplier",
        'user_count': user_count,
        'bid_count': bid_count,
        'supplier_count': supplier_count,
        'service_provider_count':service_provider_count,
        'manufacturer_count': manufacturer_count,
        'referral_count': referral_count,
        "bid_cancelled_count": bid_cancelled_count,
        "on_going_bid_count": on_going_bid_count,
        "bid_finished_count": bid_finished_count
    }


    return render(request, 'excution-team-dashboard.html',context)


@login_required(login_url='panel_login')
def service_support_dashboard(request):
    # Get the count of all users
    user_count = Users.objects.count()

    # Get the count of users based on user_type
    supplier_count = Users.objects.filter(user_type='supplier').count()
    manufacturer_count = Users.objects.filter(user_type='manufacturer').count()
    referral_count = Users.objects.filter(user_type='referral').count()
    service_provider_count = Users.objects.filter(user_type='service provider').count()

    bid_count = Bids.objects.count()
    on_going_bid_count= Bids.objects.filter(bid_status='on_going').count()
    bid_finished_count = Bids.objects.filter(bid_status='finished').count()
    bid_cancelled_count = Bids.objects.filter(bid_status='cancelled').count()


    context = {
        "manufacturer": "manufacturer",
        "supplier": "supplier",
        'user_count': user_count,
        'bid_count': bid_count,
        'supplier_count': supplier_count,
        'service_provider_count':service_provider_count,
        'manufacturer_count': manufacturer_count,
        'referral_count': referral_count,
        "bid_cancelled_count": bid_cancelled_count,
        "on_going_bid_count": on_going_bid_count,
        "bid_finished_count": bid_finished_count
    }

    return render(request, 'service-support-dashboard.html',context)


@login_required(login_url='panel_login')
def freelancers_dashboard(request):
    # Get the count of all users
    user_count = Users.objects.count()

    # Get the count of users based on user_type
    supplier_count = Users.objects.filter(user_type='supplier').count()
    manufacturer_count = Users.objects.filter(user_type='manufacturer').count()
    referral_count = Users.objects.filter(user_type='referral').count()
    service_provider_count = Users.objects.filter(user_type='service provider').count()

    bid_count = Bids.objects.count()
    on_going_bid_count= Bids.objects.filter(bid_status='on_going').count()
    bid_finished_count = Bids.objects.filter(bid_status='finished').count()
    bid_cancelled_count = Bids.objects.filter(bid_status='cancelled').count()


    context = {
        "manufacturer": "manufacturer",
        "supplier": "supplier",
        'user_count': user_count,
        'bid_count': bid_count,
        'supplier_count': supplier_count,
        'service_provider_count':service_provider_count,
        'manufacturer_count': manufacturer_count,
        'referral_count': referral_count,
        "bid_cancelled_count": bid_cancelled_count,
        "on_going_bid_count": on_going_bid_count,
        "bid_finished_count": bid_finished_count
    }

    return render(request, 'freelancers-dashboard.html',context)


@login_required(login_url='panel_login')
def referral_dashboard(request):

    # Get the count of all users
    user_count = Users.objects.count()

    # Get the count of users based on user_type
    supplier_count = Users.objects.filter(user_type='supplier').count()
    manufacturer_count = Users.objects.filter(user_type='manufacturer').count()
    referral_count = Users.objects.filter(user_type='referral').count()
    service_provider_count = Users.objects.filter(user_type='service provider').count()

    bid_count = Bids.objects.count()
    on_going_bid_count= Bids.objects.filter(bid_status='on_going').count()
    bid_finished_count = Bids.objects.filter(bid_status='finished').count()
    bid_cancelled_count = Bids.objects.filter(bid_status='cancelled').count()


    context = {
        "manufacturer": "manufacturer",
        "supplier": "supplier",
        'user_count': user_count,
        'bid_count': bid_count,
        'supplier_count': supplier_count,
        'service_provider_count':service_provider_count,
        'manufacturer_count': manufacturer_count,
        'referral_count': referral_count,
        "bid_cancelled_count": bid_cancelled_count,
        "on_going_bid_count": on_going_bid_count,
        "bid_finished_count": bid_finished_count
    }

    return render(request, 'referral-dashboard.html',context)


def my_profile(request):
    user_id = request.user.id

    user = PanelUser.objects.filter(id=user_id).first()

    if user is None:
        messages.success(request, 'User not found')
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

    form = MyprofleForm(instance=user)
    if request.method == 'POST':
        form = MyprofleForm(request.POST, instance=user)
        if form.is_valid():
            form.save()

            messages.success(request, 'Profile updated successfully')
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

    return render(request, 'my-profile.html', {'form': form, 'user_id': user_id})



def edit_user_profile(request, id):
    user = Users.objects.get(id=id)


    form = UserForm(instance=user)
    if request.method == 'POST':
        form = UserForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Profile updated successfully')
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))


    return render(request, 'user-profile.html', {'form': form,'userid': id,"user":user})


def delete_user(request, id):
    user = Users.objects.get(id=id)
    user.delete()
    # delete the user  Assessments
    assessment = Assessments.objects.filter(created_by=id).first()
    if assessment:
        assessment.delete()

    # delete the user  Bids
    # bid = Bids.objects.filter(party1_id=id).first()
    # if bid:
    #     bid.delete()

    # delete the user  Products
    product = Products.objects.filter(user_id=id).first()
    if product:
        product.delete()

    # delete the user  Services
    service = Services.objects.filter(user_id=id).first()
    if service:
        service.delete()
        
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))  


def edit_user_assessment(request, id):
    assessment = Assessments.objects.filter(created_by=id).first()
    if assessment is None:
        messages.success(request, 'Assessment not found')
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
    form = AssessmentForm(instance=assessment)
    if request.method == 'POST':
        form = AssessmentForm(request.POST, instance=assessment)
        if form.is_valid():
            form.save()
            messages.success(request, 'Assessment updated successfully')
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

    return render(request, 'user-assement.html', {'form': form, 'assessment_id': id})

def delete_assessment(request, id):
    assessment = Assessments.objects.get(id=id)
    assessment.delete()
    return redirect('assement_list')


def add_panel_user(request):
    if request.method == 'POST':
        form = PanelUserForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('panel_users_admin')  # Replace with your success URL
    else:
        form = PanelUserForm()
    
    return render(request, 'add-panel-user.html', {'form': form})

def list_panel_users(request):
    users = PanelUser.objects.all()
    names = [user.name for user in users]  # Extract the names

    context = {
        'response': 'Success',
        'names': names  
    }

    return context


def panel_users(request):
    users = PanelUser.objects.all()

    context = {
        'users': users
    }

    return render(request, 'admin-type-list.html', context)

def panel_users_admin(request):
    users = PanelUser.objects.filter(user_type='Admin')

    context = {
        'users': users
    }

    return render(request, 'admin-type-list.html', context)

def panel_users_execution(request):
    users = PanelUser.objects.filter(user_type='Execution')

    context = {
        'users': users
    }

    return render(request, 'admin-type-list.html', context)

def panel_users_service_support(request):
    users = PanelUser.objects.filter(user_type='Service_support')

    context = {
        'users': users
    }

    return render(request, 'admin-type-list.html', context)

def panel_users_freelancers(request):
    users = PanelUser.objects.filter(user_type='Freelancers')

    context = {
        'users': users
    }

    return render(request, 'admin-type-list.html', context)

def panel_users_referral(request):
    users = PanelUser.objects.filter(user_type='referral')

    context = {
        'users': users
    }

    return render(request, 'admin-type-list.html', context)



def manufacturers_list(request):
    users = Users.objects.filter(user_type='manufacturer')

    context = {
        'users': users
    }

    return render(request, 'user-type-list.html', context)




def suppliers_list(request):
    users = Users.objects.filter(user_type='supplier')

    context = {
        'users': users
    }

    return render(request, 'user-type-list.html', context)



def service_provider_list(request):
    users = Users.objects.filter(user_type='service provider')

    context = {
        'users': users
    }

    return render(request, 'user-type-list.html', context)


def referral_list(request):
    users = Users.objects.filter(user_type='referral')

    context = {
        'users': users
    }

    return render(request, 'user-type-list.html', context)



def approve_disapprove_user(request, user_id):
    try:
        user = Users.objects.get(id=user_id)
        user.is_approved = 'no' if user.is_approved == 'yes' else 'yes'
        user.save()
        # return  on same page
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
    
    except Users.DoesNotExist:
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
    



def bids_by_type_view(request):
    bid_type = request.GET.get('bid_type', '')  # Assuming bid_type is a field in your Bids model
    bids = Bids.objects.filter(bid_type=bid_type)

    return render(request, 'bids.html', {'bids': bids, 'bid_type': bid_type,})



def bids_view(request):
    bids = Bids.objects.all()

    return render(request, 'bid-list.html', {'bids': bids,})


def bids_view_one(request):
    bids = Bids.objects.filter(bid_type="one_time")

    return render(request, 'bid-list.html', {'bids': bids,})


def bids_view_real(request):
    bids = Bids.objects.filter(bid_type="real_time")

    return render(request, 'bid-list.html', {'bids': bids,})


def bid_approve_disapprove(request, bid_id):
    try:
        bid = Bids.objects.get(id=bid_id)
        bid.is_approved = 'no' if bid.is_approved == 'yes' else 'yes'
        bid.save()
        # return  on same page
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
    
    except Bids.DoesNotExist:
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))



def bids_edit_view(request,bid_id):    

    bid = Bids.objects.filter(id=bid_id)
    itemscategory=ItemsCategory.objects.all()
    itemssubcategory=ItemsSubCategories.objects.all()

    if request.method == "POST":
            bid_id = request.POST.get('id')

            bid = get_object_or_404(Bids, id=bid_id)

            bid.item = request.POST.get('item')
            bid.description = request.POST.get('description')
            bid.price = request.POST.get('price') 
            bid.start_bid_price = request.POST.get('start_bid_price') 
            bid.bid_type = request.POST.get('bid_type')
            bid.bid_win_type = request.POST.get('bid_win_type')
            bid.bid_category = request.POST.get('bid_category')
            bid.bid_sub_category = request.POST.get('bid_sub_category')
            bid.bid_opening_time = request.POST.get('bid_opening_time')
            bid.bid_closing_time = request.POST.get('bid_closing_time')
            bid.bid_status = request.POST.get('bid_status')
            bid.percentage_inc_dec = request.POST.get('percentage_inc_dec')
            bid.bid_quantity = request.POST.get('bid_quantity')

            bid.save()  

            update_current_bid(bid_id, request.POST.get('start_bid_price'))

            bid_id = request.POST.get('id')      
            bid = Bids.objects.filter(id=bid_id)
            return render(request, 'edit-bid.html', {'bid_data': bid, 'bid_id': bid_id,'itemscategory':itemscategory,'itemssubcategory':itemssubcategory})

    return render(request, 'edit-bid.html', {'bid_data': bid, 'bid_id': bid_id,'itemscategory':itemscategory,'itemssubcategory':itemssubcategory})



def bids_delete_view(request,bid_id):
    bid = Bids.objects.get(id=bid_id)
    bid.delete()
    return redirect('bids_view')



def toggle_approval(request, bid_id):
    try:

        bid = Bids.objects.get(id=bid_id)
        bid.is_approved = 'no' if bid.is_approved == 'yes' else 'yes'
        bid.save()
        return JsonResponse({'status': 'success', 'is_approved': bid.is_approved})
    except Bids.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Bid not found'}, status=404)
    


def toggle_user_approval(request, user_id):
    try:
        user = Users.objects.get(id=user_id)
        user.is_approved = 'no' if user.is_approved == 'yes' else 'yes'
        user.save()
        return JsonResponse({'status': 'success', 'is_approved': user.is_approved})
    except Bids.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Bid not found'}, status=404)
    

def toggle_user_assessment_approval(request, assessment_id):
    try:
        assessment = Assessments.objects.get(id=assessment_id)
        assessment.is_approved = 'no' if assessment.is_approved == 'yes' else 'yes'
        assessment.save()
        return JsonResponse({'status': 'success', 'is_approved': assessment.is_approved})
    except Bids.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Assessment not found'}, status=404)


def toggle_product_approval(request, product_id):
    try:
        product = Products.objects.get(id=product_id)
        product.is_approved = 'no' if product.is_approved == 'yes' else 'yes'
        product.save()
        return JsonResponse({'status': 'success', 'is_approved': product.is_approved})
    except Bids.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Product not found'}, status=404)


def units_list(request):
    categories = Unit.objects.all()

    context = {
        'units': categories
    }

    return render(request, 'unit-list.html', context)

def add_unit(request):
    form = UnitForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('units_list')
    
    return render(request, 'add-update-unit.html', {'form': form,})

def update_unit(request, units_id):
    category = Unit.objects.get(id=units_id)
    form = UnitForm(request.POST or None, instance=category)
    if form.is_valid():
        form.save()
        return redirect('units_list')
    
    return render(request, 'add-update-unit.html', {'form': form, 'category_id': units_id,})

def delete_unit(request, units_id):
    category = Unit.objects.get(id=units_id)
    category.delete()
    return redirect('units_list')



def category_list(request):
    categories = ItemsCategory.objects.all()

    context = {
        'categories': categories
    }

    return render(request, 'category-list.html', context)




def add_category(request):
    form = ItemsCategoryForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('category_list')
    
    return render(request, 'add-update-category.html', {'form': form,})


def category_approve_disapprove(request,id):
    SubCategories = ItemsCategory.objects.get(id=id)
    SubCategories.is_approved = 'no' if SubCategories.is_approved == 'yes' else 'yes'
    SubCategories.save()
    return redirect('category_list')




def update_category(request, category_id):
    category = ItemsCategory.objects.get(id=category_id)
    form = ItemsCategoryForm(request.POST or None, instance=category)
    if form.is_valid():
        form.save()
        return redirect('category_list')
    
    return render(request, 'add-update-category.html', {'form': form, 'category_id': category_id,})




def delete_category(request, category_id):
    category = ItemsCategory.objects.get(id=category_id)
    subcategory=ItemsSubCategories.objects.filter(category_id=category_id) 
    if subcategory:
        for sub in subcategory:
            sub.delete()
    category.delete()
    return redirect('category_list')



def sub_category_list(request):
    sub_categories = ItemsSubCategories.objects.all()
    # get category list from ItemsSubCategories.ItemsCategory


    for category in sub_categories:
        category_name = ItemsCategory.objects.get(id=category.category_id)
        category.category_id = category_name.category
    
    context = {
        'sub_categories': sub_categories

    }

    return render(request, 'sub-category-list.html', context)



def add_sub_category(request):
    
    form = ItemsSubCategoriesForm(request.POST or None)
    
    if form.is_valid():
        form.save()
        return redirect('sub_category_list')
    
    return render(request, 'add-update-sub-category.html', {'form': form})



def subcategory_approve_disapprove(request,id):
    SubCategories = ItemsSubCategories.objects.get(id=id)

    SubCategories.is_approved = 'no' if SubCategories.is_approved == 'yes' else 'yes'
    SubCategories.save()
    return redirect('sub_category_list')





def update_sub_category(request, sub_category_id):
    sub_category = ItemsSubCategories.objects.get(id=sub_category_id)
    form = ItemsSubCategoriesForm(request.POST or None, instance=sub_category)
    if form.is_valid(): 
        form.save()
        return redirect('sub_category_list')
    

    return render(request, 'add-update-sub-category.html', {'form': form, 'sub_category_id': sub_category_id,})



def delete_subcategory(request,id):

    sub_category = ItemsSubCategories.objects.get(id=id)
    sub_category.delete()

    return redirect('sub_category_list')



def view_product(request,user_id):
    product = Products.objects.filter(user_id=user_id)
    type = "product"
    return render(request, 'view-product-scrvices-scrap.html' , {"datas":product, "type":type})


def view_scrvices(request,user_id):
    services = Services.objects.filter(user_id=user_id)
    type = "services"
    return render(request, 'view-product-scrvices-scrap.html',{"datas":services, "type":type})


def view_scrap(request,user_id):
    scrap = Scraps.objects.filter(user_id=user_id)
    type = "scrap"
    return render(request, 'view-product-scrvices-scrap.html',{"datas":scrap, "type":type})




def refers_list(request):
    refers = Refers.objects.all()
    return render(request, 'refer-list.html',{"refers":refers})

def delete_referral(request,referral_id):
    referral = Refers.objects.get(id=referral_id)
    referral.delete()
    return redirect(request.META.get('HTTP_REFERER'))



def manufacturer_bids_list(request,user_id):
    bids = Bids.objects.filter(party1_id=user_id)
    return render(request, 'bid-list.html',{"bids":bids})




def supplier_bids_list(request,user_id):
    bids = Bids.objects.all()

    bid_list =[]

    for b in bids:

        if b.other_parties :
            data_list = json.loads(b.other_parties)

            for data in list(data_list):
                if data['party_id'] == user_id:
                    bid_list.append(b)

    return render(request, 'supplier-bid-list.html',{"bids":bid_list,"user_id":user_id})


def supplier_view_bid_details(request,user_id,bid_id):
    bid = Bids.objects.get(id=bid_id)
    for b in json.loads(bid.other_parties):
        if b['party_id'] == user_id:
            bid.other_parties = b
            break
    return render(request, 'view-bid-details.html',{"bid":bid})




def domains_list(request):
    domanis = Domains.objects.all()

    return render(request, 'user-domanis-list.html',{"datas":domanis})


def add_domains(request):
    form = DomainsForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('domains_list')
    
    return render(request, 'add-update-domanis-user-type.html', {'form': form})

def delete_domains(request,id):
    domain = Domains.objects.get(id=id)
    domain.delete()
    
    return redirect('domains_list')

def domain_approve_disapprove(request,id):
    domain = Domains.objects.get(id=id)
    domain.is_approved = 'no' if domain.is_approved == 'yes' else 'yes'
    domain.save()
    return redirect('domains_list')

def products_list(request):
    products = Products.objects.all()
    return render(request, 'product-list.html',{"datas":products})



def edit_product(request,id):
    product = Products.objects.get(id=id)
    form = ProductForm(request.POST or None, instance=product)
    if form.is_valid():
        form.save()
        return redirect('products_list')
    
    return render(request, 'edit-product.html', {'form': form, 'product_id': id,})


def delete_product(request,id):
    product = Products.objects.get(id=id)
    product.delete()
    return redirect('products_list')


def product_approve_disapprove(request,id):
    product = Products.objects.get(id=id)

    product.is_approved = 'no' if product.is_approved == 'yes' else 'yes'
    product.save()
    return redirect('products_list')


def assement_list(request):
    assements = Assessments.objects.all()

    return render(request, 'assessment-list.html',{"datas":assements})


def user_type_list(request):
    user_type = UserType.objects.all()
    types="user_type"
    return render(request, 'user-domanis-list.html',{"datas":user_type, "types":types})

def user_approve_disapprove(request,id):
    usertype = UserType.objects.get(id=id)

    usertype.is_approved = 'no' if usertype.is_approved == 'yes' else 'yes'
    usertype.save()
    return redirect('user_type_list')

def add_user_type(request):
    form = UserTypeForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('user_type_list')
    
    return render(request, 'add-update-domanis-user-type.html', {'form': form})


def banner_list(request):
    banners = Banner.objects.all()
    return render(request, 'banner-list.html',{"datas":banners})

def add_banner(request):
    form = BannerForm(request.POST or None,request.FILES)
    if form.is_valid():
        form.save()
        return redirect('banner_list')
    
    return render(request, 'add-update-banner.html', {'form': form})

def update_banner(request,id) :
    banner = Banner.objects.get(id=id)
    form = BannerForm(request.POST or None, instance=banner)
    if form.is_valid(): 
        form.save()
        return redirect('banner_list')
    
    return render(request, 'add-update-banner.html', {'form': form})



def delete_banner(request,id) :

    banner = Banner.objects.get(id=id)

    banner.delete()
    
    return redirect('banner_list')

def export_to_excel_sup(request):
    # Create an in-memory workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Data'

    # Get all the field names from the model
    model_fields = [field.name for field in Users._meta.get_fields()]
    
    # Define the headers dynamically
    sheet.append(model_fields)

    # Retrieve the data from the database
    data = Users.objects.filter(user_type="supplier").values_list(*model_fields)

    # Write the data to the sheet
    for row in data:
        sheet.append([str(value) for value in row])

    # Create a HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=data.xlsx'
    workbook.save(response)

    return response

def export_to_excel_man(request):
    # Create an in-memory workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Data'

    # Get all the field names from the model
    model_fields = [field.name for field in Users._meta.get_fields()]
    
    # Define the headers dynamically
    sheet.append(model_fields)

    # Retrieve the data from the database
    data = Users.objects.filter(user_type="manufacturers").values_list(*model_fields)

    # Write the data to the sheet
    for row in data:
        sheet.append([str(value) for value in row])

    # Create a HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=data.xlsx'
    workbook.save(response)

    return response

def export_to_excel_service(request):
    # Create an in-memory workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Data'

    # Get all the field names from the model
    model_fields = [field.name for field in Users._meta.get_fields()]
    
    # Define the headers dynamically
    sheet.append(model_fields)

    # Retrieve the data from the database
    data = Users.objects.filter(user_type="service provider").values_list(*model_fields)

    # Write the data to the sheet
    for row in data:
        sheet.append([str(value) for value in row])

    # Create a HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=data.xlsx'
    workbook.save(response)

    return response

def export_to_excel_referrals(request):
    # Create an in-memory workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Data'

    # Get all the field names from the model
    model_fields = [field.name for field in Users._meta.get_fields()]
    
    # Define the headers dynamically
    sheet.append(model_fields)

    # Retrieve the data from the database
    data = Users.objects.filter(user_type="referral").values_list(*model_fields)

    # Write the data to the sheet
    for row in data:
        sheet.append([str(value) for value in row])

    # Create a HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=data.xlsx'
    workbook.save(response)

    return response

def export_to_excel_refers(request):
    # Create an in-memory workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Data'

    # Get all the field names from the model
    model_fields = [field.name for field in Refers._meta.get_fields()]
    
    # Define the headers dynamically
    sheet.append(model_fields)

    # Retrieve the data from the database
    data = Refers.objects.all().values_list(*model_fields)

    # Write the data to the sheet
    for row in data:
        sheet.append([str(value) for value in row])

    # Create a HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=data.xlsx'
    workbook.save(response)

    return response

def export_to_excel_one_time(request):
    # Create an in-memory workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Data'

    # Get all the field names from the model
    model_fields = [field.name for field in Bids._meta.get_fields()]
    
    # Define the headers dynamically
    sheet.append(model_fields)

    # Retrieve the data from the database
    data = Bids.objects.filter(bid_type="one_time").values_list(*model_fields)

    # Write the data to the sheet
    for row in data:
        sheet.append([str(value) for value in row])

    # Create a HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=data.xlsx'
    workbook.save(response)

    return response

def export_to_excel_real_time(request):
    # Create an in-memory workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Data'

    # Get all the field names from the model
    model_fields = [field.name for field in Bids._meta.get_fields()]
    
    # Define the headers dynamically
    sheet.append(model_fields)

    # Retrieve the data from the database
    data = Bids.objects.filter(bid_type="real_time").values_list(*model_fields)

    # Write the data to the sheet
    for row in data:
        sheet.append([str(value) for value in row])

    # Create a HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=data.xlsx'
    workbook.save(response)

    return response

def export_to_excel_cat(request):
    # Create an in-memory workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Data'

    # Get all the field names from the model
    model_fields = [field.name for field in ItemsCategory._meta.get_fields()]
    
    # Define the headers dynamically
    sheet.append(model_fields)

    # Retrieve the data from the database
    data = ItemsCategory.objects.all().values_list(*model_fields)

    # Write the data to the sheet
    for row in data:
        sheet.append([str(value) for value in row])

    # Create a HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=data.xlsx'
    workbook.save(response)

    return response

def export_to_excel_sub_cat(request):
    # Create an in-memory workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Data'

    # Get all the field names from the model
    model_fields = [field.name for field in ItemsSubCategories._meta.get_fields()]
    
    # Define the headers dynamically
    sheet.append(model_fields)

    # Retrieve the data from the database
    data = ItemsSubCategories.objects.all().values_list(*model_fields)

    # Write the data to the sheet
    for row in data:
        sheet.append([str(value) for value in row])

    # Create a HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=data.xlsx'
    workbook.save(response)

    return response   

def export_to_excel_products(request):
    # Create an in-memory workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Data'

    # Get all the field names from the model
    model_fields = [field.name for field in Products._meta.get_fields()]
    
    # Define the headers dynamically
    sheet.append(model_fields)

    # Retrieve the data from the database
    data = Products.objects.all().values_list(*model_fields)

    # Write the data to the sheet
    for row in data:
        sheet.append([str(value) for value in row])

    # Create a HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=data.xlsx'
    workbook.save(response)

    return response   

def export_to_excel_user_types(request):
    # Create an in-memory workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Data'

    # Get all the field names from the model
    model_fields = [field.name for field in UserType._meta.get_fields()]
    
    # Define the headers dynamically
    sheet.append(model_fields)

    # Retrieve the data from the database
    data = UserType.objects.all().values_list(*model_fields)

    # Write the data to the sheet
    for row in data:
        sheet.append([str(value) for value in row])

    # Create a HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=data.xlsx'
    workbook.save(response)

    return response   

def export_to_excel_user_domains(request):
    # Create an in-memory workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Data'

    # Get all the field names from the model
    model_fields = [field.name for field in Domains._meta.get_fields()]
    
    # Define the headers dynamically
    sheet.append(model_fields)

    # Retrieve the data from the database
    data = Domains.objects.all().values_list(*model_fields)

    # Write the data to the sheet
    for row in data:
        sheet.append([str(value) for value in row])

    # Create a HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=data.xlsx'
    workbook.save(response)

    return response   